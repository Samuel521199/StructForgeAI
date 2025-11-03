"""
Excel解析器
"""
from pathlib import Path
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl import load_workbook

from data_parser.base_parser import BaseParser
from core.logging_config import logger


class ExcelParser(BaseParser):
    """Excel文件解析器（.xlsx格式）"""
    
    def __init__(self):
        pass
    
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """解析Excel文件"""
        try:
            workbook = load_workbook(str(file_path), data_only=True)
            
            result = {
                "format": "excel",
                "sheets": {}
            }
            
            # 解析每个Sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._parse_sheet(sheet)
                result["sheets"][sheet_name] = sheet_data
            
            result["sheet_count"] = len(workbook.sheetnames)
            result["active_sheet"] = workbook.active.title
            
            return result
            
        except Exception as e:
            logger.error(f"Excel解析失败: {e}")
            raise
    
    def _parse_sheet(self, sheet) -> Dict[str, Any]:
        """解析单个Sheet"""
        rows = []
        headers = None
        
        # 读取所有数据
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 跳过空行
            if not any(cell for cell in row):
                continue
            
            # 第一行作为表头
            if row_idx == 1:
                headers = [str(cell) if cell is not None else f"Column{i+1}" 
                          for i, cell in enumerate(row)]
                continue
            
            # 构建行数据
            row_data = {}
            for col_idx, cell_value in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else f"Column{col_idx+1}"
                row_data[header] = cell_value if cell_value is not None else ""
            
            rows.append(row_data)
        
        # 推断数据类型
        field_types = {}
        if rows and headers:
            field_types = self._infer_field_types(rows)
        
        return {
            "headers": headers or [],
            "rows": rows,
            "row_count": len(rows),
            "field_types": field_types
        }
    
    def _infer_field_types(self, rows: List[Dict]) -> Dict[str, str]:
        """推断字段类型（与CSV类似）"""
        if not rows:
            return {}
        
        field_types = {}
        
        for field_name in rows[0].keys():
            values = [row.get(field_name) for row in rows[:100]]
            
            types_found = set()
            for value in values:
                if value is None or value == "":
                    continue
                if isinstance(value, (int, float)):
                    if isinstance(value, int):
                        types_found.add("integer")
                    else:
                        types_found.add("number")
                elif isinstance(value, str):
                    types_found.add("string")
                elif isinstance(value, bool):
                    types_found.add("boolean")
            
            if "integer" in types_found and "number" not in types_found:
                field_types[field_name] = "integer"
            elif "number" in types_found:
                field_types[field_name] = "number"
            elif "boolean" in types_found:
                field_types[field_name] = "boolean"
            else:
                field_types[field_name] = "string"
        
        return field_types
    
    def validate(self, data: Dict[str, Any]) -> bool:
        """验证Excel数据"""
        return "format" in data and "sheets" in data
    
    def export(self, data: Dict[str, Any], output_path: Path) -> bool:
        """导出为Excel文件"""
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            workbook = openpyxl.Workbook()
            
            sheets = data.get("sheets", {})
            
            # 如果有多个Sheet，删除默认Sheet
            if len(sheets) > 1:
                workbook.remove(workbook.active)
            
            # 创建每个Sheet
            for sheet_name, sheet_data in sheets.items():
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(title=sheet_name)
                
                # 写入表头
                headers = sheet_data.get("headers", [])
                if headers:
                    sheet.append(headers)
                
                # 写入数据
                rows = sheet_data.get("rows", [])
                for row_data in rows:
                    row_values = [row_data.get(header, "") for header in headers]
                    sheet.append(row_values)
            
            workbook.save(str(output_path))
            return True
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return False
    
    def detect_schema(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """检测Excel结构Schema"""
        schema = {
            "type": "spreadsheet",
            "format": "excel",
            "sheets": {}
        }
        
        sheets = data.get("sheets", {})
        
        for sheet_name, sheet_data in sheets.items():
            sheet_schema = {
                "columns": {}
            }
            
            headers = sheet_data.get("headers", [])
            field_types = sheet_data.get("field_types", {})
            
            for header in headers:
                sheet_schema["columns"][header] = {
                    "type": field_types.get(header, "string"),
                    "position": headers.index(header)
                }
            
            schema["sheets"][sheet_name] = sheet_schema
        
        return schema

